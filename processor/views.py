import re
import shutil

import openpyxl
import os

import pandas as pd
import json
import requests
from django.shortcuts import render, redirect, get_object_or_404
from django.http import FileResponse, Http404
from .forms import UploadFileForm
from .models import UploadedFile
from django.conf import settings
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet


def upload_and_list_files(request):
    """Загрузка одного файла и отображение состояния."""
    current_file = UploadedFile.objects.filter(is_current=True).first()
    processed_filename = None

    # 🧼 Чистим, если файл не существует
    if current_file and not os.path.exists(current_file.file.path):
        current_file.delete()
        current_file = None

    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            UploadedFile.objects.filter(is_current=True).delete()
            new_file = form.save(commit=False)
            new_file.is_current = True
            new_file.save()
            return redirect("home")
        # 🔻 если form невалидна — мы просто "падаем" дальше и отрисовываем шаблон с ошибками
    else:
        form = UploadFileForm()

    # Проверяем финальный файл
    if current_file:
        base_name = os.path.splitext(os.path.basename(current_file.file.name))[0]
        final_name = f"{base_name}_final.xlsx"
        final_path = os.path.join(settings.BASE_DIR, "uploads", final_name)
        if os.path.exists(final_path):
            processed_filename = final_name

    chatgpt_table = request.session.pop("chatgpt_table", None)

    return render(request, "processor/index.html", {
        "form": form,
        "current_file": current_file,
        "processed_filename": processed_filename,
        "chatgpt_table": chatgpt_table,  # 👈 добавили
    })


def upload_file(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect("file_list")  # Перенаправляем на список загруженных файлов
    else:
        form = UploadFileForm()
    return render(request, "processor/upload.html", {"form": form})


def process_xlsx(file_path):
    """
    Загружает Excel-файл, удаляет дубликаты по колонке '№',
    удаляет полностью пустые строки и столбцы,
    и сохраняет новый файл.
    """
    # Читаем файл
    df = pd.read_excel(file_path)

    # Удаляем дубликаты по колонке "№", оставляя только первую строку
    if "№" in df.columns:
        df = df.drop_duplicates(subset="№", keep="first")

    columns_to_remove = ["Шаг", "Ожидаемый результат", "Примечания/вопросы", "Примечания/вопросы (аналитик)"]
    df = df.drop(columns=[col for col in columns_to_remove if col in df.columns])

    # Сохраняем новый файл
    new_file_path = file_path.replace(".xlsx", "_processed.xlsx")
    df.to_excel(new_file_path, index=False)

    return new_file_path


def file_list(request):
    """Выводит только актуальные загруженные файлы + обработанные."""

    # Фильтрация: показываем только те записи, у которых файл существует физически
    all_uploaded = UploadedFile.objects.all()
    existing_uploaded_files = [
        f for f in all_uploaded if os.path.exists(f.file.path)
    ]

    # Обработанные файлы из папки /uploads/
    uploads_dir = os.path.join(settings.BASE_DIR, "uploads")
    processed_files = []
    if os.path.exists(uploads_dir):
        for filename in os.listdir(uploads_dir):
            if "_final" in filename and filename.endswith(".xlsx"):
                processed_files.append(filename)

    return render(request, "processor/file_list.html", {
        "files": existing_uploaded_files,
        "processed_files": processed_files,
    })


def convert_xlsx_to_text(file_path):
    """Читает .xlsx, конвертирует в JSON (текстовый формат)."""
    df = pd.read_excel(file_path)
    return df.to_json(orient="records", force_ascii=False, indent=4)  # JSON с сохранением символов


def convert_text_to_xlsx(json_text, output_path):
    """Принимает текст, извлекает JSON-массив и сохраняет его в .xlsx."""
    try:
        # Преобразуем в DataFrame
        data = json.loads(json_text)
        df = pd.DataFrame(data)
        df.to_excel(output_path, index=False)
        return True

    except Exception as e:
        print(f"Ошибка преобразования: {e}")
        return None


def apply_priorities_from_chatgpt(original_path, chatgpt_path):
    """Добавляет колонку Приоритет и сохраняет структуру вложенных строк."""
    original_df = pd.read_excel(original_path)
    chatgpt_df = pd.read_excel(chatgpt_path)

    if "№" not in original_df.columns or "Ответы" not in chatgpt_df.columns:
        print("Ошибка: отсутствует нужная колонка в одном из файлов.")
        return None

    # Мапим приоритеты
    priority_mapping = chatgpt_df.set_index("№")["Ответы"].to_dict()
    original_df["Приоритет"] = original_df["№"].map(priority_mapping)

    # Загружаем исходный файл с группировкой
    wb = load_workbook(original_path)
    ws: Worksheet = wb.active

    # Ищем свободную колонку (например, следующую после последней)
    gpt_col_idx = ws.max_column + 1
    ws.cell(row=1, column=gpt_col_idx, value="Приоритет")

    # Строим маппинг "№" → приоритет
    num_col_index = None
    for col_idx, cell in enumerate(ws[1], 1):
        if str(cell.value).strip() == "№":
            num_col_index = col_idx
            break

    if not num_col_index:
        print("Не найдена колонка '№'")
        return

    # Применяем приоритет к каждой строке
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row_num_cell = row[num_col_index - 1]
        row_num_value = row_num_cell.value
        priority = priority_mapping.get(row_num_value)
        if priority is not None:
            ws.cell(row=row_num_cell.row, column=gpt_col_idx, value=priority)

    # Сохраняем новый файл, сохранив группировку
    final_path = original_path.replace(".xlsx", "_final.xlsx")
    wb.save(final_path)
    return final_path


def send_to_chatgpt(text_data, prompt):
    """Отправляет текст в ChatGPT API и получает обработанный JSON."""
    url = "https://api.proxyapi.ru/openai/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {os.getenv('PROXYAPI_KEY')}",
        "Content-Type": "application/json",
    }

    messages = [
        {"role": "developer", "content": f"{prompt}"},
        {"role": "user", "content": f"Данные:\n{text_data}"}
    ]

    payload = {
        "model": "gpt-4o-mini",
        "messages": messages,
    }

    response = requests.post(url, headers=headers, json=payload)

    if response.status_code == 200:
        text = response.json()["choices"][0]["message"]["content"]
        # Извлекаем массив из текста — от первой [ до последней ]
        match = re.search(r'\[.*]', text, re.DOTALL)
        if not match:
            return text

        cleaned_json = match.group(0)

        return cleaned_json

    return None


def process_with_chatgpt(request):
    """Обрабатывает текущий файл, сохраняет результат и обновляет страницу."""
    uploaded_file = UploadedFile.objects.filter(is_current=True).first()
    if not uploaded_file:
        return render(request, "processor/error.html", {"message": "Нет текущего файла для обработки."})

    original_file_path = uploaded_file.file.path
    processed_file_path = process_xlsx(original_file_path)

    text_data = convert_xlsx_to_text(processed_file_path)
    prompt = settings.CHATGPT_PROMPT
    processed_text = send_to_chatgpt(text_data, prompt)

    if processed_text:
        chatgpt_path = original_file_path.replace(".xlsx", "_chatgpt.xlsx")
        result = convert_text_to_xlsx(processed_text, chatgpt_path)

        if result:
            apply_priorities_from_chatgpt(original_file_path, chatgpt_path)
            try:
                parsed_data = json.loads(processed_text)

                # Сначала найдём все уникальные приоритеты
                priorities = set()
                for row in parsed_data:
                    if isinstance(row, dict):
                        p = row.get("Ответы")
                        if isinstance(p, int):
                            priorities.add(p)

                # Вычисляем на сколько уменьшать
                priority_shift = 0
                if 1 not in priorities:
                    if 2 not in priorities:
                        priority_shift = 2  # ни 1, ни 2
                    else:
                        priority_shift = 1  # нет 1, но есть 2

                # Применяем сдвиг
                if priority_shift > 0:
                    for row in parsed_data:
                        if isinstance(row, dict) and isinstance(row.get("Ответы"), int):
                            row["Ответы"] = max(1, row["Ответы"] - priority_shift)

                # Сохраняем в сессию
                request.session['chatgpt_table'] = parsed_data

            except Exception as e:
                print("Ошибка парсинга ответа от ChatGPT:", e)

    return redirect("home")
